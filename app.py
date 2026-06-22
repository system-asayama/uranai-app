"""四柱推命（日柱）60干支占いアプリ ＋ ログイン認証システム。

- 生年月日から日柱の干支（六十干支）を算出する占い機能
- セッションベースの認証
- パスワードはハッシュ化して保存
- admin / user のロールによるアクセス制御
- 管理者はユーザー一覧・作成・ロール変更・削除が可能
"""
import os
from datetime import date
from functools import wraps
from io import BytesIO

from flask import (
    Flask,
    Response,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)

import sangokushi
import shichu
from models import ROLE_ADMIN, ROLE_USER, ROLES, User, db


def _normalize_db_url(url: str) -> str:
    # SQLAlchemy は postgres:// を認識しないため postgresql:// に変換
    if url.startswith("postgres://"):
        return url.replace("postgres://", "postgresql://", 1)
    return url


def create_app() -> Flask:
    app = Flask(__name__)
    app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-change-me")

    database_url = os.environ.get("DATABASE_URL")
    if database_url:
        app.config["SQLALCHEMY_DATABASE_URI"] = _normalize_db_url(database_url)
    else:
        # DATABASE_URL が無い場合はローカル SQLite にフォールバック
        app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///app.db"
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

    # 相性占いのログイン要否（一時的に公開中。"1"/"true" でログイン必須に戻す）
    app.config["AISHO_REQUIRE_LOGIN"] = (
        os.environ.get("AISHO_REQUIRE_LOGIN", "0").lower() in ("1", "true", "yes")
    )

    db.init_app(app)

    with app.app_context():
        db.create_all()
        _seed_admin()

    _register_routes(app)
    return app


def _seed_admin() -> None:
    """初期管理者アカウントを作成する（既に存在する場合は何もしない）。"""
    admin_username = os.environ.get("ADMIN_USERNAME", "admin")
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")

    if User.query.filter_by(username=admin_username).first() is None:
        admin = User(username=admin_username, role=ROLE_ADMIN)
        admin.set_password(admin_password)
        db.session.add(admin)
        db.session.commit()


# ---------------------------------------------------------------------------
# 認証ヘルパー
# ---------------------------------------------------------------------------
def current_user():
    user_id = session.get("user_id")
    if user_id is None:
        return None
    return db.session.get(User, user_id)


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if current_user() is None:
            flash("ログインが必要です。", "error")
            return redirect(url_for("login"))
        return view(*args, **kwargs)

    return wrapped


def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        user = current_user()
        if user is None:
            flash("管理者ログインが必要です。", "error")
            return redirect(url_for("admin_login"))
        if not user.is_admin:
            flash("管理者権限が必要です。", "error")
            return redirect(url_for("dashboard"))
        return view(*args, **kwargs)

    return wrapped


# ---------------------------------------------------------------------------
# ルーティング
# ---------------------------------------------------------------------------
def _register_routes(app: Flask) -> None:
    @app.context_processor
    def inject_user():
        return {
            "current_user": current_user(),
            "aisho_requires_login": app.config["AISHO_REQUIRE_LOGIN"],
        }

    @app.context_processor
    def inject_helpers():
        def char_url(index):
            """キャラ画像URL。実画像があれば更新時刻を、無ければデータ版を
            バージョンに付与し、差し替え時にキャッシュを確実に更新させる。"""
            chars_dir = os.path.join(app.root_path, "static", "characters")
            ver = "svg"
            for ext in ("png", "jpg", "jpeg", "webp", "svg"):
                path = os.path.join(chars_dir, f"{index}.{ext}")
                if os.path.exists(path):
                    ver = str(int(os.path.getmtime(path)))
                    break
            return url_for("char_image", index=index, v=ver)

        return {"char_url": char_url}

    @app.route("/")
    def index():
        # トップは誰でも使える四柱推命占いページ
        return redirect(url_for("uranai"))

    @app.route("/uranai", methods=["GET", "POST"])
    def uranai():
        """生年月日（と時刻）から四柱（命式）を算出し鑑定する（ログイン不要）。"""
        result = None
        form = {"year": "", "month": "", "day": "", "hour": "", "minute": ""}

        if request.method == "POST":
            for key in form:
                form[key] = (request.form.get(key) or "").strip()

            try:
                birth = date(int(form["year"]), int(form["month"]), int(form["day"]))
            except (ValueError, TypeError):
                flash("正しい生年月日を入力してください。", "error")
                return render_template("uranai.html", result=result, form=form)

            # 時刻は任意。時が空なら時刻不明として時柱を省略する
            hour = minute = None
            if form["hour"] != "":
                try:
                    hour = int(form["hour"])
                    minute = int(form["minute"]) if form["minute"] != "" else 0
                    if not (0 <= hour <= 23 and 0 <= minute <= 59):
                        raise ValueError
                except (ValueError, TypeError):
                    flash("時刻は 0〜23 時・0〜59 分で入力してください。", "error")
                    return render_template("uranai.html", result=result, form=form)

            if birth.year < 1873:
                # 日本がグレゴリオ暦を採用した1873年以降のみ対応
                flash("1873年以降の日付を入力してください。", "error")
            elif birth > date.today():
                flash("未来の日付は占えません。", "error")
            else:
                result = shichu.compute_four_pillars(birth, hour, minute or 0)

        character = (sangokushi.get_character(result.day_index)
                     if result is not None else None)
        return render_template(
            "uranai.html", result=result, form=form, character=character
        )

    @app.route("/uranai/list")
    def uranai_list():
        """六十干支と三国志キャラの一覧（早見表）。"""
        rows = list(zip(shichu.all_fortunes(), sangokushi.all_characters()))
        return render_template("uranai_list.html", rows=rows)

    @app.route("/team", methods=["GET", "POST"])
    def team():
        """複数メンバーの五行からチームビルディングを分析する（ログイン不要）。"""
        slots = 6
        result = None
        form = {f"m{i}_{k}": "" for i in range(slots)
                for k in ("name", "year", "month", "day")}

        if request.method == "POST":
            for key in form:
                form[key] = (request.form.get(key) or "").strip()

            people = []
            error = False
            for i in range(slots):
                y = form[f"m{i}_year"]
                mo = form[f"m{i}_month"]
                d = form[f"m{i}_day"]
                if not (y or mo or d):
                    continue  # 空行はスキップ
                label = form[f"m{i}_name"]
                who = label or f"メンバー{i + 1}"
                try:
                    birth = date(int(y), int(mo), int(d))
                except (ValueError, TypeError):
                    flash(f"{who}の生年月日を正しく入力してください。", "error")
                    error = True
                    break
                if birth.year < 1873 or birth > date.today():
                    flash(f"{who}は1873年以降〜今日までの日付で入力してください。",
                          "error")
                    error = True
                    break
                people.append((label, shichu.compute_four_pillars(birth)))

            if not error and len(people) < 2:
                flash("メンバーは2人以上入力してください。", "error")
            elif not error:
                result = shichu.analyze_team(people)

        return render_template("team.html", result=result, form=form,
                               slots=slots, element_color=shichu.ELEMENT_COLOR,
                               element_team=shichu.ELEMENT_TEAM)

    @app.route("/biorhythm", methods=["GET", "POST"])
    def biorhythm():
        """十干（通変星）によるバイオリズム分析（日/月/年・ログイン不要）。"""
        result = None
        form = {"year": "", "month": "", "day": "",
                "s_year": "", "s_month": "", "s_day": "", "mode": "day"}

        if request.method == "POST":
            for key in form:
                form[key] = (request.form.get(key) or "").strip()
            mode = form["mode"] if form["mode"] in shichu.BIO_MODES else "day"
            form["mode"] = mode
            try:
                birth = date(int(form["year"]), int(form["month"]), int(form["day"]))
            except (ValueError, TypeError):
                flash("生年月日を正しく入力してください。", "error")
                return render_template("biorhythm.html", result=result, form=form)

            # 起点日（任意）。未入力なら今日
            start = date.today()
            if form["s_year"] or form["s_month"] or form["s_day"]:
                try:
                    start = date(int(form["s_year"]), int(form["s_month"]),
                                 int(form["s_day"]))
                except (ValueError, TypeError):
                    flash("起点日を正しく入力してください。", "error")
                    return render_template("biorhythm.html", result=result, form=form)

            if birth.year < 1873 or birth > date.today():
                flash("生年月日は1873年以降〜今日までで入力してください。", "error")
            else:
                result = shichu.biorhythm(birth, start, mode=mode)

        return render_template("biorhythm.html", result=result, form=form,
                               modes=shichu.BIO_MODES)

    @app.route("/char/<int:index>")
    def char_image(index):
        """三国志キャラの画像。static/characters/<index>.<ext> があればそれを、
        無ければ生成した SVG アバターを返す（ログイン不要）。"""
        if not 0 <= index < 60:
            return ("not found", 404)

        chars_dir = os.path.join(app.root_path, "static", "characters")
        for ext in ("png", "jpg", "jpeg", "webp", "svg"):
            if os.path.exists(os.path.join(chars_dir, f"{index}.{ext}")):
                return send_from_directory(chars_dir, f"{index}.{ext}")

        svg = sangokushi.character_svg(
            sangokushi.get_character(index), shichu.ganzhi_name(index)
        )
        return Response(svg, mimetype="image/svg+xml",
                        headers={"Cache-Control": "public, max-age=86400"})

    @app.route("/characters.xlsx")
    def characters_xlsx():
        """60キャラクターを軍勢・名前付きの Excel でダウンロード（ログイン不要）。"""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "三国志キャラ一覧"

        headers = ["No.", "干支", "名前", "読み", "軍勢", "肩書き"]
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="1A1A2E")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 軍勢→名前の順に並べる
        faction_order = {"魏": 0, "蜀": 1, "呉": 2, "群雄": 3}
        chars = sorted(
            sangokushi.all_characters(),
            key=lambda c: (faction_order.get(c.faction, 9), c.index),
        )

        for c in chars:
            ws.append([
                c.index + 1,
                shichu.ganzhi_name(c.index),
                c.name,
                c.yomi,
                c.faction,
                c.title,
            ])
            # 軍勢セルを勢力色で塗る
            row = ws.max_row
            fill = PatternFill("solid", fgColor=c.color.lstrip("#").upper())
            faction_cell = ws.cell(row=row, column=5)
            faction_cell.fill = fill
            faction_cell.font = Font(bold=True, color="FFFFFF")
            faction_cell.alignment = Alignment(horizontal="center")

        widths = [6, 8, 14, 18, 8, 28]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = "A2"

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument."
                     "spreadsheetml.sheet",
            as_attachment=True,
            download_name="sangokushi_characters.xlsx",
        )

    @app.route("/aisho", methods=["GET", "POST"])
    def aisho():
        """二人の生年月日から相性を占う。

        AISHO_REQUIRE_LOGIN が True のときのみログイン必須
        （現在は一時的に未ログインでも利用可）。
        """
        if app.config["AISHO_REQUIRE_LOGIN"] and current_user() is None:
            flash("ログインが必要です。", "error")
            return redirect(url_for("login"))
        result = None
        # a_* = あなた / b_* = お相手
        form = {f"{who}_{k}": "" for who in ("a", "b")
                for k in ("year", "month", "day", "hour", "minute")}

        if request.method == "POST":
            for key in form:
                form[key] = (request.form.get(key) or "").strip()

            people = []
            error = False
            for who, label in (("a", "あなた"), ("b", "お相手")):
                try:
                    birth = date(
                        int(form[f"{who}_year"]),
                        int(form[f"{who}_month"]),
                        int(form[f"{who}_day"]),
                    )
                except (ValueError, TypeError):
                    flash(f"{label}の生年月日を正しく入力してください。", "error")
                    error = True
                    break

                hour = minute = None
                if form[f"{who}_hour"] != "":
                    try:
                        hour = int(form[f"{who}_hour"])
                        minute = (int(form[f"{who}_minute"])
                                  if form[f"{who}_minute"] != "" else 0)
                        if not (0 <= hour <= 23 and 0 <= minute <= 59):
                            raise ValueError
                    except (ValueError, TypeError):
                        flash(f"{label}の時刻は 0〜23 時・0〜59 分で入力してください。",
                              "error")
                        error = True
                        break

                if birth.year < 1873:
                    flash(f"{label}は 1873 年以降の日付を入力してください。", "error")
                    error = True
                    break
                if birth > date.today():
                    flash(f"{label}に未来の日付は使えません。", "error")
                    error = True
                    break

                people.append(shichu.compute_four_pillars(birth, hour, minute or 0))

            if not error:
                fp_a, fp_b = people
                result = {
                    "a": fp_a,
                    "b": fp_b,
                    "char_a": sangokushi.get_character(fp_a.day_index),
                    "char_b": sangokushi.get_character(fp_b.day_index),
                    "compat": shichu.compatibility(fp_a, fp_b),
                }

        return render_template("aisho.html", result=result, form=form)

    @app.route("/register", methods=["GET", "POST"])
    def register():
        """利用者（user ロール）の新規登録。"""
        if request.method == "POST":
            username = (request.form.get("username") or "").strip()
            password = request.form.get("password") or ""
            confirm = request.form.get("confirm") or ""

            if not username or not password:
                flash("ユーザー名とパスワードを入力してください。", "error")
            elif password != confirm:
                flash("パスワードが一致しません。", "error")
            elif User.query.filter_by(username=username).first() is not None:
                flash("そのユーザー名は既に使われています。", "error")
            else:
                user = User(username=username, role=ROLE_USER)
                user.set_password(password)
                db.session.add(user)
                db.session.commit()
                flash("登録が完了しました。ログインしてください。", "success")
                return redirect(url_for("login"))

        return render_template("register.html")

    @app.route("/login", methods=["GET", "POST"])
    def login():
        """利用者用ログインページ。"""
        if current_user() is not None:
            return redirect(url_for("dashboard"))

        if request.method == "POST":
            username = (request.form.get("username") or "").strip()
            password = request.form.get("password") or ""

            user = User.query.filter_by(username=username).first()
            if user is not None and user.check_password(password):
                if user.is_admin:
                    # 管理者は管理者用ログインを使う
                    flash("管理者は管理者ログインページからログインしてください。", "error")
                    return redirect(url_for("admin_login"))
                session.clear()
                session["user_id"] = user.id
                flash(f"ようこそ、{user.username} さん。", "success")
                return redirect(url_for("dashboard"))

            flash("ユーザー名またはパスワードが正しくありません。", "error")

        return render_template("login.html")

    @app.route("/admin/login", methods=["GET", "POST"])
    def admin_login():
        """管理者用ログインページ。"""
        user = current_user()
        if user is not None:
            return redirect(url_for("admin_users" if user.is_admin else "dashboard"))

        if request.method == "POST":
            username = (request.form.get("username") or "").strip()
            password = request.form.get("password") or ""

            user = User.query.filter_by(username=username).first()
            if user is not None and user.check_password(password):
                if not user.is_admin:
                    # 一般利用者はこのページからログインできない
                    flash("このページは管理者専用です。利用者ログインをご利用ください。", "error")
                    return redirect(url_for("login"))
                session.clear()
                session["user_id"] = user.id
                flash(f"管理者としてログインしました（{user.username}）。", "success")
                return redirect(url_for("admin_users"))

            flash("ユーザー名またはパスワードが正しくありません。", "error")

        return render_template("admin_login.html")

    @app.route("/logout")
    def logout():
        was_admin = (current_user() or None) and current_user().is_admin
        session.clear()
        flash("ログアウトしました。", "success")
        return redirect(url_for("admin_login" if was_admin else "login"))

    @app.route("/dashboard")
    @login_required
    def dashboard():
        return render_template("dashboard.html", user=current_user())

    # --- 管理者専用 ------------------------------------------------------
    @app.route("/admin/users")
    @admin_required
    def admin_users():
        users = User.query.order_by(User.created_at.asc()).all()
        return render_template("admin_users.html", users=users, roles=ROLES)

    @app.route("/admin/users/create", methods=["POST"])
    @admin_required
    def admin_create_user():
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        role = request.form.get("role") or ROLE_USER

        if role not in ROLES:
            role = ROLE_USER

        if not username or not password:
            flash("ユーザー名とパスワードを入力してください。", "error")
        elif User.query.filter_by(username=username).first() is not None:
            flash("そのユーザー名は既に使われています。", "error")
        else:
            user = User(username=username, role=role)
            user.set_password(password)
            db.session.add(user)
            db.session.commit()
            flash(f"ユーザー「{username}」を作成しました。", "success")

        return redirect(url_for("admin_users"))

    @app.route("/admin/users/<int:user_id>/role", methods=["POST"])
    @admin_required
    def admin_update_role(user_id):
        user = db.session.get(User, user_id)
        if user is None:
            flash("ユーザーが見つかりません。", "error")
            return redirect(url_for("admin_users"))

        new_role = request.form.get("role")
        if new_role not in ROLES:
            flash("無効なロールです。", "error")
            return redirect(url_for("admin_users"))

        # 最後の管理者を降格させないよう保護
        if user.is_admin and new_role != ROLE_ADMIN and _admin_count() <= 1:
            flash("最後の管理者の権限は変更できません。", "error")
            return redirect(url_for("admin_users"))

        user.role = new_role
        db.session.commit()
        flash(f"「{user.username}」のロールを {new_role} に変更しました。", "success")
        return redirect(url_for("admin_users"))

    @app.route("/admin/users/<int:user_id>/delete", methods=["POST"])
    @admin_required
    def admin_delete_user(user_id):
        user = db.session.get(User, user_id)
        if user is None:
            flash("ユーザーが見つかりません。", "error")
            return redirect(url_for("admin_users"))

        if user.id == current_user().id:
            flash("自分自身は削除できません。", "error")
            return redirect(url_for("admin_users"))

        if user.is_admin and _admin_count() <= 1:
            flash("最後の管理者は削除できません。", "error")
            return redirect(url_for("admin_users"))

        db.session.delete(user)
        db.session.commit()
        flash(f"ユーザー「{user.username}」を削除しました。", "success")
        return redirect(url_for("admin_users"))


def _admin_count() -> int:
    return User.query.filter_by(role=ROLE_ADMIN).count()


app = create_app()


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000, debug=True)
